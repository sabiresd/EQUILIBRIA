"""
Agent Simulation (WF-1) — GridBalance / Defi 10.

Remplace l'appel a une API meteo temps reel par un tirage dans le dataset
horaire NASA POWER (CERES/MERRA2) de l'annee 2023.

Deux modes :
  spot   : une heure tiree au hasard          -> pour un flux "temps reel"
  window : N heures contigues depuis un debut -> ce que l'Agent Calcul attend

La sortie est un dict JSON-serialisable conforme a contracts/schemas.json
(WF2Request : correlation_id, series[], battery, tariffs).

Usage CLI :
    python agent_simulation.py                      # spot, 1 heure aleatoire
    python agent_simulation.py --mode window -n 360 # serie de 360 h
    python agent_simulation.py --seed 42            # tirage reproductible
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import random
import uuid
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from facture_onee import (
    FACTURE_FILE,
    Facture,
    FactureError,
    baseline_cost,
    calibration_factors,
    facture_to_contract,
    load_facture,
)

# --------------------------------------------------------------------------
# Configuration — tout ce qui est "metier" est ici, rien n'est en dur plus bas.
# --------------------------------------------------------------------------

DATA_FILE = Path(__file__).with_name("Data2023.xlsx")  # .csv accepte aussi

# Horizon par defaut, aligne sur le contrat WF2Request (15 jours).
DEFAULT_HORIZON_HOURS = 360

# Le dataset NASA POWER est horodate en UTC ; la tarification ONEE s'applique
# a l'heure locale marocaine (UTC+1). Sans ce decalage, les heures de pointe
# sont decalees d'une heure.
UTC_OFFSET_HOURS = 1

# Tarification ONEE (MAD/kWh) — valeurs de DEMONSTRATION, non officielles ANRE.
# Bornes en heure LOCALE, intervalle [debut, fin[.
TARIFF_PERIODS: list[tuple[int, int, str, float]] = [
    (22, 6, "creuse", 0.65),   # 22h -> 06h (traverse minuit)
    (6, 18, "normale", 0.90),  # 06h -> 18h
    (18, 22, "pointe", 1.40),  # 18h -> 22h
]

# Profil de consommation de l'usine (MW).
DEMAND_BASE_MW = 0.8   # nuit, machines a l'arret
DEMAND_PEAK_MW = 2.2   # plein regime
DEMAND_RAMP_START = 6  # debut de montee en charge (heure locale)
DEMAND_RAMP_END = 22   # retour au regime de nuit
DEMAND_NOISE = 0.03    # +/- 3 % de bruit, pour ne pas avoir un profil trop lisse

# --- Parc eolien ---
# Dimensionne sur la facture (~1.3 MW moyen), et non sur les 6-15 MW de l'ancien
# stub : un parc surdimensionne couvre tout, la batterie ne se vide jamais, et il
# ne reste aucun deficit a resoudre.
N_WT = 2                # nombre de turbines
P_RATED_MW = 0.65       # puissance nominale d'une turbine
ETA_WT = 0.95           # rendement de conversion (generatrice + convertisseur)

WIND_CUT_IN_MS = 3.0    # sous ce vent, la turbine ne produit rien
WIND_RATED_MS = 12.0    # au-dessus, elle est a puissance nominale
WIND_CUT_OUT_MS = 25.0  # au-dessus, arret de securite

WIND_HUB_HEIGHT_M = 80.0  # les formules prennent V a la NACELLE ; WS10M est a 10 m
WIND_SHEAR_ALPHA = 0.14   # exposant de cisaillement (terrain degage)

# Puissance maximale du parc : N_wt x P_rated x eta_WT (formule 7).
WIND_CAPACITY_MW = N_WT * P_RATED_MW * ETA_WT

SOLAR_CAPACITY_MW = 1.0
SOLAR_PERF_RATIO = 0.80  # pertes onduleur, salissure, cablage
SOLAR_TEMP_COEF = -0.004  # -0.4 %/degre au-dessus de 25 C
SOLAR_STC_IRRADIANCE = 1000.0  # W/m2

# Batterie, transmise telle quelle a l'Agent Calcul. Dimensionnee pour ecreter
# les 4 h de pointe du soir d'un site a ~1.8 MW de pointe (4 h x 2 MW = 8 MWh),
# pas les 40 MWh / 10 MW de l'ancienne demo, qui donnaient 31 h d'autonomie et
# rendaient le probleme trivial.
BATTERY = {
    "capacity_mwh": 8.0,
    "p_max_mw": 2.0,
    "soc_min": 0.10,
    "efficiency": 0.92,
    "degradation_cost_mwh": 120.0,
}

# Facteur de puissance : la facture donne des kVA, le calcul raisonne en MW.
COS_PHI = 0.8

COLUMNS = ["YEAR", "MO", "DY", "HR", "irr", "T2M", "RH2M", "WS10M", "WD10M"]

DISCLAIMER = (
    "Donnees de simulation. Non connecte aux systemes de l'ONEE. "
    "Tarifs affiches a titre de demonstration, non officiels ANRE."
)


class DatasetError(RuntimeError):
    """Le dataset est introuvable, illisible, ou ne contient aucune ligne valide."""


# --------------------------------------------------------------------------
# Chargement
# --------------------------------------------------------------------------


@dataclass
class WeatherRow:
    year: int
    month: int
    day: int
    hour_utc: int
    irr: float      # W/m2
    t2m: float      # degres C
    rh2m: float     # %
    ws10m: float    # m/s a 10 m
    wd10m: float    # degres

    @property
    def timestamp_utc(self) -> datetime:
        return datetime(self.year, self.month, self.day, self.hour_utc, tzinfo=timezone.utc)


def _to_float(value: Any) -> float | None:
    """NASA POWER stocke ses nombres en texte et code les manquants par -999."""
    if value is None:
        return None
    try:
        f = float(str(value).strip().replace(",", "."))
    except (TypeError, ValueError):
        return None
    if math.isnan(f) or f <= -900:  # sentinelle de valeur manquante
        return None
    return f


def _rows_from_csv(path: Path) -> list[dict[str, Any]]:
    with path.open(newline="", encoding="utf-8-sig") as fh:
        return list(csv.DictReader(fh))


def _rows_from_excel(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise DatasetError(
            f"Lecture de {path.name} impossible : openpyxl n'est pas installe. "
            "Faites `pip install openpyxl`, ou exportez le fichier en .csv."
        ) from exc

    ws = load_workbook(path, read_only=True, data_only=True).active
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    return [dict(zip(header, r)) for r in rows]


def load_dataset(path: Path | str = DATA_FILE) -> list[WeatherRow]:
    """Charge le dataset, en ne gardant que les lignes exploitables.

    Le fichier NASA embarque un bloc d'en-tete et une colonne de commentaires :
    toute ligne dont une colonne utile n'est pas numerique est ecartee, sans
    faire echouer le chargement.
    """
    path = Path(path)
    if not path.exists():
        raise DatasetError(
            f"Dataset introuvable : {path}\n"
            f"Placez le fichier NASA POWER 2023 a cet emplacement, ou passez "
            f"--data / le parametre `data_path`."
        )

    suffix = path.suffix.lower()
    if suffix == ".csv":
        raw = _rows_from_csv(path)
    elif suffix in (".xlsx", ".xlsm"):
        raw = _rows_from_excel(path)
    else:
        raise DatasetError(f"Extension non geree : {suffix} (attendu .csv ou .xlsx)")

    if raw and not all(c in raw[0] for c in COLUMNS):
        missing = [c for c in COLUMNS if c not in (raw[0] or {})]
        raise DatasetError(
            f"Colonnes manquantes dans {path.name} : {missing}. "
            f"Attendu : {COLUMNS}"
        )

    rows: list[WeatherRow] = []
    for r in raw:
        vals = {c: _to_float(r.get(c)) for c in COLUMNS}
        if any(v is None for v in vals.values()):
            continue  # ligne d'en-tete, de commentaire, ou trouee
        rows.append(
            WeatherRow(
                year=int(vals["YEAR"]),
                month=int(vals["MO"]),
                day=int(vals["DY"]),
                hour_utc=int(vals["HR"]),
                irr=vals["irr"],
                t2m=vals["T2M"],
                rh2m=vals["RH2M"],
                ws10m=vals["WS10M"],
                wd10m=vals["WD10M"],
            )
        )

    if not rows:
        raise DatasetError(
            f"{path.name} a ete lu mais ne contient aucune ligne exploitable "
            f"(colonnes {COLUMNS} toutes numeriques)."
        )

    rows.sort(key=lambda w: (w.year, w.month, w.day, w.hour_utc))
    return rows


# --------------------------------------------------------------------------
# Modeles : tarif, consommation, production
# --------------------------------------------------------------------------


def local_hour(row: WeatherRow, offset_h: int = UTC_OFFSET_HOURS) -> int:
    return (row.timestamp_utc + timedelta(hours=offset_h)).hour


def tariff_for_hour(hour_local: int) -> tuple[str, float]:
    """Retourne (periode, prix MAD/kWh) pour une heure locale donnee."""
    for start, end, label, price in TARIFF_PERIODS:
        in_period = start <= hour_local < end if start < end else (hour_local >= start or hour_local < end)
        if in_period:
            return label, price
    raise ValueError(f"Heure hors de toute periode tarifaire : {hour_local}")


def demand_mw(hour_local: int, rng: random.Random | None = None) -> float:
    """Consommation de l'usine : plateau de nuit, montee en cloche en journee.

    Le profil est une demi-sinusoide entre DEMAND_RAMP_START et DEMAND_RAMP_END,
    ce qui evite les marches d'escalier d'un simple if/else et donne a l'Agent
    Calcul une courbe derivable.

    rng=None : profil NU, sans bruit. C'est cette version qu'on utilise pour
    calibrer sur la facture — sinon le facteur d'echelle bougerait a chaque tirage.
    """
    if DEMAND_RAMP_START <= hour_local < DEMAND_RAMP_END:
        span = DEMAND_RAMP_END - DEMAND_RAMP_START
        phase = (hour_local - DEMAND_RAMP_START) / span
        shape = math.sin(math.pi * phase)
        value = DEMAND_BASE_MW + (DEMAND_PEAK_MW - DEMAND_BASE_MW) * shape
    else:
        value = DEMAND_BASE_MW

    if rng is not None:
        value *= 1.0 + rng.uniform(-DEMAND_NOISE, DEMAND_NOISE)
    return round(max(0.0, value), 3)


def profile_kwh_by_period(hours: int) -> dict[str, float]:
    """Energie du profil SYNTHETIQUE nu, par poste horaire, sur `hours` heures.

    Sert de denominateur au recalage : on compare ce que le profil consommerait
    a ce que la facture dit avoir ete consomme, sur la meme duree.
    """
    totals = {"creuse": 0.0, "normale": 0.0, "pointe": 0.0}
    for h in range(hours):
        hl = h % 24
        period, _ = tariff_for_hour(hl)
        totals[period] += demand_mw(hl) * 1000.0  # MW sur 1 h -> kWh
    return totals


def wind_speed_at_hub(ws10m: float) -> float:
    """Extrapole la vitesse de 10 m a la hauteur de nacelle (loi de puissance).

    Les formules (6)-(8) prennent V a la NACELLE. La colonne WS10M du dataset NASA
    est mesuree a 10 m. Sans cette extrapolation, on sous-estime la production
    d'environ 30 % a 80 m de haut.

        V_hub = V_10 x (H_hub / 10) ^ alpha
    """
    return ws10m * (WIND_HUB_HEIGHT_M / 10.0) ** WIND_SHEAR_ALPHA


def wind_power_mw(ws10m: float) -> float:
    """Puissance du parc eolien (MW), formules (6) a (8).

        (6)  P = 0                                             si V < V_cut_in OU V > V_cut_out
        (7)  P = N_wt x P_rated x eta_WT                       si V_rated <= V <= V_cut_out
        (8)  P = N_wt x P_rated x (V^3 - V_ci^3)
                                 -------------- x eta_WT       si V_cut_in <= V < V_rated
                                 (V_r^3 - V_ci^3)

    Note : la formule (6) telle qu'ecrite dit "V < V_cut_in ET V > V_cut_out",
    ce qu'aucune vitesse ne peut satisfaire. C'est un OU — coquille du papier.
    """
    v = wind_speed_at_hub(ws10m)

    # (6) hors plage de fonctionnement : pas assez de vent, ou arret de securite.
    if v < WIND_CUT_IN_MS or v > WIND_CUT_OUT_MS:
        return 0.0

    # (7) plateau : la turbine est bridee a sa puissance nominale.
    if v >= WIND_RATED_MS:
        return round(N_WT * P_RATED_MW * ETA_WT, 3)

    # (8) zone cubique : la puissance du vent varie comme le cube de sa vitesse.
    ratio = (v**3 - WIND_CUT_IN_MS**3) / (WIND_RATED_MS**3 - WIND_CUT_IN_MS**3)
    return round(N_WT * P_RATED_MW * ratio * ETA_WT, 3)


def solar_power_mw(irr: float, t2m: float) -> float:
    """Production PV, derating thermique inclus (le rendement chute avec T2M)."""
    if irr <= 0:
        return 0.0
    cell_temp = t2m + 25.0 * (irr / SOLAR_STC_IRRADIANCE)  # approximation NOCT
    derate = 1.0 + SOLAR_TEMP_COEF * max(0.0, cell_temp - 25.0)
    power = SOLAR_CAPACITY_MW * (irr / SOLAR_STC_IRRADIANCE) * SOLAR_PERF_RATIO * derate
    return round(min(max(power, 0.0), SOLAR_CAPACITY_MW), 3)


# --------------------------------------------------------------------------
# Construction du point de serie
# --------------------------------------------------------------------------


def _anchor_index(
    rows: list[WeatherRow],
    anchor: str,
    rng: random.Random,
    offset_h: int,
) -> int:
    """Index de depart de la fenetre.

    anchor="now"    : l'heure REELLE d'envoi. Le dataset couvre 2023 : on projette
                      la date du jour sur le meme creneau (mois, jour, heure) de 2023.
                      La serie envoyee a 15 h demarre donc a 15 h.
    anchor="random" : tirage aleatoire (demo reproductible avec --seed).
    """
    if anchor == "random":
        return rng.randrange(len(rows))

    if anchor != "now":
        raise ValueError(f"anchor inconnu : {anchor!r} (attendu 'now' ou 'random')")

    now_utc = datetime.now(timezone.utc)
    exact = next(
        (
            i
            for i, r in enumerate(rows)
            if (r.month, r.day, r.hour_utc) == (now_utc.month, now_utc.day, now_utc.hour)
        ),
        None,
    )
    if exact is not None:
        return exact

    # Creneau absent du dataset (29 fevrier, heure trouee...) : on prend le plus
    # proche dans l'annee plutot que d'echouer.
    def distance(r: WeatherRow) -> int:
        return abs(
            (r.month * 31 * 24 + r.day * 24 + r.hour_utc)
            - (now_utc.month * 31 * 24 + now_utc.day * 24 + now_utc.hour)
        )

    return min(range(len(rows)), key=lambda i: distance(rows[i]))


def build_point(
    row: WeatherRow,
    h_index: int,
    rng: random.Random,
    offset_h: int,
    calibration: dict[str, float] | None = None,
    start_hour_local: int = 0,
) -> dict[str, Any]:
    """Un point de serie, STRICTEMENT conforme au contrat SeriesPoint.

    Le contrat est `extra="forbid"` : tout champ supplementaire (timestamp,
    temperature, periode tarifaire...) ferait rejeter le payload par le backend.
    Ces informations vivent dans `meta`, pas ici.

    L'heure locale est celle que WF-2 recalculera : (h + start_hour_local) % 24.
    On utilise la MEME formule ici, pour que la conso recalee et le cout calcule
    en aval portent sur la meme periode tarifaire. Deriver l'heure de `row` a la
    place marcherait aussi, mais divergerait silencieusement de WF-2 le jour ou
    la serie aurait un trou.
    """
    hl = (h_index + start_hour_local) % 24
    period, _ = tariff_for_hour(hl)
    demand = demand_mw(hl, rng)

    if calibration:  # recalage sur la facture, poste par poste
        demand = round(demand * calibration.get(period, 1.0), 3)

    return {
        "h": h_index,
        "wind_ms": round(row.ws10m, 2),
        "ghi": round(row.irr, 1),
        "prod_wind_mw": wind_power_mw(row.ws10m),
        "prod_solar_mw": solar_power_mw(row.irr, row.t2m),
        "demand_mw": demand,
    }


def simulate(
    data_path: Path | str = DATA_FILE,
    mode: str = "spot",
    hours: int = DEFAULT_HORIZON_HOURS,
    seed: int | None = None,
    correlation_id: str | None = None,
    utc_offset: int = UTC_OFFSET_HOURS,
    facture_path: Path | str | None = FACTURE_FILE,
    anchor: str = "now",
) -> dict[str, Any]:
    """Point d'entree unique — c'est cette fonction que le noeud Python appelle.

    mode="spot"   : un seul point (l'heure courante).
    mode="window" : `hours` heures contigues a partir de l'ancrage.

    anchor="now"    : la serie demarre a l'heure REELLE d'envoi. C'est cette heure
                      que WF-2 utilisera (via start_hour_local), donc l'heure a
                      laquelle vous envoyez est bien celle qu'il calcule.
    anchor="random" : tirage aleatoire, reproductible avec --seed. Pour les demos.

    facture_path=None desactive le recalage : la demande reste synthetique et les
    tarifs restent ceux de TARIFF_PERIODS.
    """
    if mode not in ("spot", "window"):
        raise ValueError(f"mode inconnu : {mode!r} (attendu 'spot' ou 'window')")

    rows = load_dataset(data_path)
    rng = random.Random(seed)
    cid = correlation_id or str(uuid.uuid4())

    # --- 1. La facture, si elle est fournie : tarifs + recalage + baseline ---
    facture: Facture | None = None
    calibration: dict[str, float] | None = None
    if facture_path is not None:
        facture = load_facture(facture_path)
        # On compare le profil nu et la facture sur la MEME duree : celle facturee.
        calibration = calibration_factors(
            profile_kwh_by_period(facture.heures_facturees), facture
        )

    # --- 2. Selection de la fenetre ---
    if hours <= 0:
        raise ValueError("hours doit etre > 0")
    if hours > len(rows):
        raise DatasetError(
            f"Fenetre de {hours} h demandee, mais le dataset n'a que {len(rows)} h."
        )

    start = _anchor_index(rows, anchor, rng, utc_offset)
    length = 1 if mode == "spot" else hours
    # Enroulement en fin d'annee : une fenetre ancree au 25 decembre doit pouvoir
    # deborder sur janvier plutot que d'echouer.
    selected = [rows[(start + k) % len(rows)] for k in range(length)]

    # L'heure locale du point h n'est pas h % 24 : elle depend de l'heure a
    # laquelle la serie DEMARRE. On la transmet a WF-2, qui en deduit la periode
    # tarifaire. Sans elle, une serie envoyee a 15 h verrait sa pointe facturee
    # au tarif creuse.
    start_hour_local = local_hour(selected[0], utc_offset)

    series = [
        build_point(r, i, rng, utc_offset, calibration, start_hour_local)
        for i, r in enumerate(selected)
    ]

    # --- 3. Tarifs : ceux de la facture s'il y en a une, sinon ceux par defaut ---
    if facture is not None:
        tariffs = facture.tariffs_mad_mwh()
        tariffs_kwh = dict(facture.prix_mad_kwh)
    else:
        tariffs = {label: round(price * 1000, 2) for _, _, label, price in TARIFF_PERIODS}
        tariffs_kwh = {label: price for _, _, label, price in TARIFF_PERIODS}

    payload: dict[str, Any] = {
        # --- contrat WF2Request : injectable tel quel dans l'Agent Calcul ---
        "correlation_id": cid,
        "series": series,
        "battery": dict(BATTERY),
        "tariffs": tariffs,  # MAD/MWh (x1000 vs le MAD/kWh de la facture)
        "start_hour_local": start_hour_local,
        "meta": {
            "source": f"NASA POWER CERES/MERRA2 - {Path(data_path).name}",
            "mode": mode,
            "anchor": anchor,
            "hours": len(series),
            "seed": seed,
            "utc_offset_hours": utc_offset,
            "sent_at_local": datetime.now(timezone.utc)
            .astimezone(timezone(timedelta(hours=utc_offset)))
            .isoformat(),
            "start_hour_local": start_hour_local,
            "period_start_utc": selected[0].timestamp_utc.isoformat(),
            "period_end_utc": selected[-1].timestamp_utc.isoformat(),
            "tariffs_mad_kwh": tariffs_kwh,
            "disclaimer": DISCLAIMER,
        },
    }

    # --- 4. Facture + baseline : le "avant", que WF-2 comparera a l'"apres" ---
    if facture is not None:
        demand_kwh = {"creuse": 0.0, "normale": 0.0, "pointe": 0.0}
        for point in series:
            # Meme formule que WF-2 : la baseline doit ventiler l'energie sur les
            # memes postes horaires que le cout auquel elle sera comparee.
            period, _ = tariff_for_hour((point["h"] + start_hour_local) % 24)
            demand_kwh[period] += point["demand_mw"] * 1000.0  # MW sur 1 h -> kWh

        payload["facture"] = facture_to_contract(facture)
        payload["baseline"] = baseline_cost(demand_kwh, facture, len(series))
        # Le plafond de soutirage EST la puissance souscrite. C'est lui qui cree le
        # deficit : les jours sans vent, la production s'effondre, la batterie se
        # vide, et le site tape dans le plafond qu'il a lui-meme sous-souscrit.
        payload["grid_cap_mw"] = round(facture.puissance_souscrite_kva * COS_PHI / 1000, 3)
        payload["meta"]["calibration_factors"] = calibration
        payload["meta"]["cos_phi"] = COS_PHI
        if facture.raw.get("_warnings"):
            payload["meta"]["facture_warnings"] = facture.raw["_warnings"]

    return payload


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------


def main() -> int:
    p = argparse.ArgumentParser(description="Agent Simulation (WF-1) — tirage NASA POWER 2023")
    p.add_argument("--data", default=str(DATA_FILE), help="chemin du .csv ou .xlsx")
    p.add_argument("--mode", choices=("spot", "window"), default="window")
    p.add_argument("-n", "--hours", type=int, default=DEFAULT_HORIZON_HOURS,
                   help="taille de la fenetre (mode window)")
    p.add_argument("--anchor", choices=("now", "random"), default="now",
                   help="now = heure reelle d'envoi (defaut) ; random = tirage aleatoire")
    p.add_argument("--seed", type=int, default=None, help="graine (avec --anchor random)")
    p.add_argument("--correlation-id", default=None)
    p.add_argument("--utc-offset", type=int, default=UTC_OFFSET_HOURS)
    p.add_argument("--facture", default=str(FACTURE_FILE),
                   help="facture ONEE (JSON) : recale la conso, fournit les tarifs, etablit la baseline")
    p.add_argument("--sans-facture", action="store_true",
                   help="ignorer la facture : conso synthetique et tarifs par defaut")
    p.add_argument("-o", "--out", default=None, help="ecrire le JSON dans un fichier")
    args = p.parse_args()

    try:
        result = simulate(
            data_path=args.data,
            mode=args.mode,
            hours=args.hours,
            seed=args.seed,
            correlation_id=args.correlation_id,
            utc_offset=args.utc_offset,
            facture_path=None if args.sans_facture else args.facture,
            anchor=args.anchor,
        )
    except (DatasetError, FactureError) as exc:
        kind = "facture" if isinstance(exc, FactureError) else "dataset"
        print(json.dumps({"error": kind, "detail": str(exc)}, ensure_ascii=False, indent=2))
        return 1
    except ValueError as exc:
        print(json.dumps({"error": "parametre", "detail": str(exc)}, ensure_ascii=False, indent=2))
        return 2

    payload = json.dumps(result, ensure_ascii=False, indent=2)
    if args.out:
        Path(args.out).write_text(payload, encoding="utf-8")
        print(f"Ecrit dans {args.out} ({len(result['series'])} points)")
    else:
        print(payload)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
